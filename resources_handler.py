# -*- coding: utf-8 -*-
"""
三防系统 - 应急资源管理模块
4类资源CRUD + Excel批量导入 + 模板生成 + 区域层级管理
"""

import json
import os
import uuid
from datetime import datetime
from typing import Dict, List, Optional, Tuple

# ==================== 配置 ====================

DATA_DIR = os.path.dirname(os.path.abspath(__file__))

RESOURCE_TYPES = {
    "personnel": {
        "file": "resources_personnel.json",
        "key": "personnel",
        "label": "人员队伍",
        "id_prefix": "P",
        "fields": ["name", "type", "team", "phone", "skills", "status", "location", "lat", "lng", "region_code", "region_name"],
        "required": ["name", "phone"],
        "excel_columns": ["姓名", "类型", "队伍", "联系电话", "技能(逗号分隔)", "状态", "所在位置", "纬度", "经度", "所属区域编码", "所属区域"],
    },
    "materials": {
        "file": "resources_materials.json",
        "key": "materials",
        "label": "物资装备",
        "id_prefix": "M",
        "fields": ["name", "category", "quantity", "unit", "location", "lat", "lng", "status", "specs", "manager", "manager_phone", "region_code", "region_name"],
        "required": ["name", "quantity"],
        "excel_columns": ["物资名称", "类别", "数量", "单位", "存放位置", "纬度", "经度", "状态", "规格说明", "管理员", "联系电话", "所属区域编码", "所属区域"],
    },
    "facilities": {
        "file": "resources_facilities.json",
        "key": "facilities",
        "label": "场所设施",
        "id_prefix": "F",
        "fields": ["name", "type", "capacity", "address", "lat", "lng", "contact", "phone", "status", "region_code", "region_name"],
        "required": ["name", "address"],
        "excel_columns": ["设施名称", "类型", "容纳人数", "详细地址", "纬度", "经度", "负责人", "联系电话", "状态", "所属区域编码", "所属区域"],
    },
    "vehicles": {
        "file": "resources_vehicles.json",
        "key": "vehicles",
        "label": "车辆运力",
        "id_prefix": "V",
        "fields": ["plate_number", "type", "model", "driver", "driver_phone", "status", "location", "lat", "lng", "region_code", "region_name"],
        "required": ["plate_number", "type"],
        "excel_columns": ["车牌号", "车辆类型", "车辆型号", "驾驶员", "联系电话", "状态", "存放位置", "纬度", "经度", "所属区域编码", "所属区域"],
    },
}

# 资源子类选项（用于前端下拉）
RESOURCE_SUBTYPES = {
    "personnel": {
        "types": ["消防队员", "水务抢险队", "应急救援队", "民兵预备役", "志愿者", "巡查人员", "值班人员", "医疗救护", "其他"],
        "teams": ["消防救援大队", "水务抢险队", "区应急救援队", "街道应急分队", "民兵应急连", "志愿救援队", "蓝天救援队", "红十字会", "其他"],
    },
    "materials": {
        "categories": ["排水设备", "发电照明", "防护用品", "救生器材", "通信设备", "帐篷棉被", "食品饮水",
                        "沙袋挡板", "抽水机泵", "发改储备物资", "民政救灾物资", "水务专用物资", "其他"],
    },
    "facilities": {
        "types": ["应急避护场所", "物资储备库", "指挥中心", "医疗救护点", "临时安置点", "应急水源", "直升机停机坪", "其他"],
    },
    "vehicles": {
        "types": ["消防车", "抢险车", "运输车", "指挥车", "救护车", "排涝车", "无人机", "冲锋舟", "其他"],
    },
}

# 字段中英文映射
FIELD_MAP = {
    "personnel": {
        "姓名": "name", "类型": "type", "队伍": "team", "联系电话": "phone",
        "技能(逗号分隔)": "skills", "状态": "status", "所在位置": "location",
        "纬度": "lat", "经度": "lng",
        "所属区域编码": "region_code", "所属区域": "region_name",
    },
    "materials": {
        "物资名称": "name", "类别": "category", "数量": "quantity", "单位": "unit",
        "存放位置": "location", "纬度": "lat", "经度": "lng",
        "状态": "status", "规格说明": "specs",
        "管理员": "manager", "联系电话": "manager_phone",
        "所属区域编码": "region_code", "所属区域": "region_name",
    },
    "facilities": {
        "设施名称": "name", "类型": "type", "容纳人数": "capacity", "详细地址": "address",
        "纬度": "lat", "经度": "lng", "负责人": "contact", "联系电话": "phone", "状态": "status",
        "所属区域编码": "region_code", "所属区域": "region_name",
    },
    "vehicles": {
        "车牌号": "plate_number", "车辆类型": "type", "车辆型号": "model",
        "驾驶员": "driver", "联系电话": "driver_phone", "状态": "status", "存放位置": "location",
        "纬度": "lat", "经度": "lng",
        "所属区域编码": "region_code", "所属区域": "region_name",
    },
}


# ==================== 工具函数 ====================

def _gen_id(prefix: str) -> str:
    return f"{prefix}{uuid.uuid4().hex[:8].upper()}"


def _data_path(filename: str) -> str:
    return os.path.join(DATA_DIR, filename)


def _load_data(resource_type: str) -> List[Dict]:
    cfg = RESOURCE_TYPES[resource_type]
    path = _data_path(cfg["file"])
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get(cfg["key"], [])
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def _save_data(resource_type: str, items: List[Dict]):
    cfg = RESOURCE_TYPES[resource_type]
    path = _data_path(cfg["file"])
    with open(path, "w", encoding="utf-8") as f:
        json.dump({cfg["key"]: items}, f, ensure_ascii=False, indent=2)


def _compute_statistics(resource_type: str, items: List[Dict]) -> Dict:
    total = len(items)
    if resource_type == "personnel":
        on_duty = sum(1 for i in items if i.get("status") == "在岗")
        standby = sum(1 for i in items if i.get("status") == "待命")
        return {"总数": total, "在岗": on_duty, "待命": standby}
    elif resource_type == "materials":
        available = sum(1 for i in items if i.get("status", "可用") == "可用")
        return {"总数": total, "可用": available, "使用中": total - available}
    elif resource_type == "facilities":
        by_type = {}
        for i in items:
            t = i.get("type", "其他")
            by_type[t] = by_type.get(t, 0) + 1
        return {"总数": total, "分类": by_type}
    elif resource_type == "vehicles":
        available = sum(1 for i in items if i.get("status", "可用") == "可用")
        return {"总数": total, "可用": available, "出勤中": total - available}
    return {"总数": total}


# ==================== CRUD ====================

def list_resources(resource_type: str, region_code: str = None) -> Dict:
    items = _load_data(resource_type)
    if region_code:
        items = [i for i in items if str(i.get("region_code", "")).startswith(str(region_code))]
    stats = _compute_statistics(resource_type, items)
    return {"items": items, "statistics": stats}


def get_resource(resource_type: str, resource_id: str) -> Optional[Dict]:
    items = _load_data(resource_type)
    for item in items:
        if item.get("id") == resource_id:
            return item
    return None


def add_resource(resource_type: str, data: Dict) -> Dict:
    cfg = RESOURCE_TYPES[resource_type]

    # 校验必填
    for field in cfg["required"]:
        if not data.get(field):
            raise ValueError(f"缺少必填字段: {field}")

    items = _load_data(resource_type)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    item = {"id": _gen_id(cfg["id_prefix"]), "create_time": now, "last_update": now}
    for field in cfg["fields"]:
        val = data.get(field, "")
        # skills 字段转列表
        if field == "skills" and isinstance(val, str):
            val = [s.strip() for s in val.split(",") if s.strip()] if val else []
        # 数值字段
        if field in ("quantity", "capacity", "lat", "lng"):
            try:
                val = float(val) if val else 0
                if field in ("quantity", "capacity"):
                    val = int(val)
            except (ValueError, TypeError):
                val = 0
        item[field] = val

    # 默认状态
    if not item.get("status"):
        item["status"] = "待命" if resource_type == "personnel" else "可用"

    items.append(item)
    _save_data(resource_type, items)
    return item


def update_resource(resource_type: str, resource_id: str, data: Dict) -> Optional[Dict]:
    cfg = RESOURCE_TYPES[resource_type]
    items = _load_data(resource_type)

    for item in items:
        if item.get("id") == resource_id:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for field in cfg["fields"]:
                if field in data:
                    val = data[field]
                    if field == "skills" and isinstance(val, str):
                        val = [s.strip() for s in val.split(",") if s.strip()] if val else []
                    if field in ("quantity", "capacity", "lat", "lng"):
                        try:
                            val = float(val) if val else 0
                            if field in ("quantity", "capacity"):
                                val = int(val)
                        except (ValueError, TypeError):
                            val = 0
                    item[field] = val
            item["last_update"] = now
            _save_data(resource_type, items)
            return item

    return None


def delete_resource(resource_type: str, resource_id: str) -> bool:
    items = _load_data(resource_type)
    new_items = [i for i in items if i.get("id") != resource_id]
    if len(new_items) == len(items):
        return False
    _save_data(resource_type, new_items)
    return True


# ==================== Excel 导入 ====================

def import_from_excel(resource_type: str, file_stream) -> Dict:
    """
    从 Excel 文件批量导入资源
    返回: {"imported": N, "failed": N, "errors": [...]}
    """
    import openpyxl

    cfg = RESOURCE_TYPES[resource_type]
    field_map = FIELD_MAP[resource_type]

    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "failed": 0, "errors": [{"row": 0, "reason": "文件为空"}]}

    # 解析表头
    header = [str(c).strip() if c else "" for c in rows[0]]

    imported = []
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            data = {}
            for col_idx, cell_val in enumerate(row):
                if col_idx >= len(header):
                    break
                col_name = header[col_idx]
                eng_field = field_map.get(col_name)
                if eng_field:
                    data[eng_field] = str(cell_val).strip() if cell_val is not None else ""

            # 跳过全空行
            if not any(data.values()):
                continue

            # 校验必填
            missing = [f for f in cfg["required"] if not data.get(f)]
            if missing:
                errors.append({"row": row_idx, "reason": f"缺少必填字段: {', '.join(missing)}"})
                continue

            item = add_resource(resource_type, data)
            imported.append(item)
        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})

    wb.close()
    return {"imported": len(imported), "failed": len(errors), "errors": errors}


def import_from_csv(resource_type: str, file_stream) -> Dict:
    """
    从 CSV 文件批量导入资源
    """
    import csv
    import io

    cfg = RESOURCE_TYPES[resource_type]
    field_map = FIELD_MAP[resource_type]

    content = file_stream.read()
    if isinstance(content, bytes):
        # 尝试 utf-8, 降级 gbk
        for enc in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return {"imported": 0, "failed": 0, "errors": [{"row": 0, "reason": "文件编码无法识别"}]}
    else:
        text = content

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"imported": 0, "failed": 0, "errors": [{"row": 0, "reason": "文件为空"}]}

    header = [c.strip() for c in rows[0]]
    imported = []
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            data = {}
            for col_idx, cell_val in enumerate(row):
                if col_idx >= len(header):
                    break
                col_name = header[col_idx]
                eng_field = field_map.get(col_name)
                if eng_field:
                    data[eng_field] = cell_val.strip() if cell_val else ""

            if not any(data.values()):
                continue

            missing = [f for f in cfg["required"] if not data.get(f)]
            if missing:
                errors.append({"row": row_idx, "reason": f"缺少必填字段: {', '.join(missing)}"})
                continue

            item = add_resource(resource_type, data)
            imported.append(item)
        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})

    return {"imported": len(imported), "failed": len(errors), "errors": errors}


# ==================== Excel 模板生成 ====================

def generate_template(resource_type: str) -> bytes:
    """
    生成 Excel 导入模板，返回 bytes
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    cfg = RESOURCE_TYPES[resource_type]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg["label"]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    columns = cfg["excel_columns"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) * 2 + 4, 12)

    # 示例数据
    examples = _get_example_row(resource_type)
    if examples:
        for col_idx, val in enumerate(examples, start=1):
            ws.cell(row=2, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _get_example_row(resource_type: str) -> list:
    if resource_type == "personnel":
        return ["张三", "消防队员", "消防救援大队", "13800138000", "水上救援,医疗急救", "待命", "XX消防站", "440112", "黄埔区"]
    elif resource_type == "materials":
        return ["移动水泵", "排水设备", 10, "台", "XX仓库", "可用", "功率200kW", "李四", "13900139000", "440112", "黄埔区"]
    elif resource_type == "facilities":
        return ["XX应急避护场所", "应急避护场所", 500, "XX区XX街道XX路XX号", 23.12, 113.26, "王五", "13700137000", "备用", "440112", "黄埔区"]
    elif resource_type == "vehicles":
        return ["粤A12345", "消防车", "东风应急车", "赵六", "13600136000", "可用", "XX停车场", "440112", "黄埔区"]
    return []


# ==================== 资源统计总览 ====================

def get_all_statistics(region_code: str = None) -> Dict:
    result = {}
    for rtype in RESOURCE_TYPES:
        items = _load_data(rtype)
        if region_code:
            items = [i for i in items if str(i.get("region_code", "")).startswith(str(region_code))]
        result[rtype] = {
            "label": RESOURCE_TYPES[rtype]["label"],
            "count": len(items),
            "statistics": _compute_statistics(rtype, items),
        }
    total = sum(r["count"] for r in result.values())
    return {"total": total, "resources": result}


def get_subtypes() -> Dict:
    """返回资源子类型选项供前端使用"""
    return RESOURCE_SUBTYPES
